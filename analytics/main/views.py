import json
from typing import List, Dict
import pandas as pd
import requests
import concurrent.futures
import openpyxl
from django.shortcuts import render
from jinja2 import FileSystemLoader
from xlsx2html import xlsx2html


def index(request):
    return render(request, 'main/home.html')

def insert_table(file_name, sheet_name):

    wb = openpyxl.load_workbook('C:/Users/anony/app/analytics/main/static/img/report.xlsx')
    worksheet = wb["Статистика по годам"]
    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)

def demand(request):
    vac_name = "web-разработчик"
    wb = openpyxl.load_workbook('C:/Users/anony/app/analytics/main/static/img/report.xlsx')
    worksheet = wb["Статистика по годам"]
    print(worksheet)

    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return render(request, 'main/demand.html', {'name_vacancy': vac_name, "excel_data": excel_data})

def geography(request):
    vac_name = "web-разработчик"
    wb = openpyxl.load_workbook('C:/Users/anony/app/analytics/main/static/img/report.xlsx')
    worksheet = wb["Статистика по городам"]
    print(worksheet)

    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return render(request, 'main/geography.html', {'name_vacancy': vac_name, "excel_data": excel_data})

def skills(request):
    vac_name = "web-разработчик"
    wb = openpyxl.load_workbook('C:/Users/anony/app/analytics/main/static/img/skills_stat.xlsx')
    worksheet = wb["Навыки"]

    excel_data = list()
    for row in worksheet.iter_rows():
        row_data = list()
        for cell in row:
            row_data.append(str(cell.value))
        excel_data.append(row_data)
    return render(request, 'main/skills.html', {'name_vacancy': vac_name, "excel_data": excel_data})

def latestVacancies(request):
    url = "https://api.hh.ru/vacancies"

    def execute_vacancies(vacancies: List[Dict[str, str]] or List[Dict[Dict[str, str], str]]) -> (List[List[str]]):
        """ формирует дату для создания итогового датафрейма
        оставляет только нужные поля: название, город, вилка оклада, дата публиукации вакансии

        Args:
            vacancies: список словарей выгруженных api содержащие информацию по вакасии

        Returns: список вакансий c зарплатами

        """
        return [
            [
                vacancy['id'],
                vacancy["name"],
                vacancy["employer"]["name"],
                vacancy["salary"]["from"],
                vacancy["salary"]["to"],
                vacancy["salary"]["currency"],
                vacancy["area"]["name"],
                vacancy["published_at"],
            ]
            for vacancy in vacancies
            if vacancy["salary"]
        ]

    def get_vacancies_items(params: dict) -> (List[Dict[str, str]] or List[Dict[Dict[str, str], str]]):
        """ поулчает json данные о вакансии по заданынм параметрам

        Args:
            params: параметры для запроса

        Returns: данные по вакансиям

        """
        return requests.get(url, params).json()["items"]


    pd.set_option('expand_frame_repr', False)
    params = dict(
        specialization=1,
        date_from="2022-12-26T00:00:00",
        date_to="2022-12-26T12:00:00",
        per_page=30,
        page=1,
    )
    res = requests.get(url, params).json()['items']
    res = execute_vacancies(res)
    for i, el in enumerate(res):
        new_get = requests.get(f"https://api.hh.ru/vacancies/{el[0]}").json()
        el.pop(0)

        el.append(new_get['description'])
        if (len(new_get['key_skills']) != 0):
            el.append(", ".join([x['name'] for x in new_get['key_skills']]))
        else:
            res.pop(i)
    res = [x for x in res if len(x) == 9]
    df = pd.DataFrame(
        res,
        columns=[
            "name",

            "company",

            "salary_from",
            "salary_to",
            "salary_currency",
            "area_name",
            "published_at",
            "discription",
            "skills",
        ],
    )
    df["published_at"] = df["published_at"].apply(lambda x: x[:10])
    df['salary'] = df['salary_from'] + df["salary_to"] / 2
    df.loc[df["salary_to"].isna(), ['salary']] = df["salary_from"]
    df = df.loc[df['salary_currency'] == 'RUR']

    df = df.drop(["salary_from", "salary_to"], axis=1)
    df = df.head(10)
    json_records = df.reset_index().to_json(orient='records')
    data = []
    data = json.loads(json_records)

    context = {"data": data}

    return render(request, 'main/latest-vacancies.html', context=context)

